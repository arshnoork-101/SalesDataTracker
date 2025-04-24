import os
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Ensure 'data' directory exists
os.makedirs("data", exist_ok=True)

# STEP 1: Get store data from user input
def get_store_data():
    stores = {}
    while True:
        store_name = input("Enter Store Name (or type 'exit' to finish): ")
        if store_name.lower() == 'exit':
            break
        store_data = []
        while True:
            product_name = input(f"Enter product name for {store_name} (or type 'done' to stop adding products): ")
            if product_name.lower() == 'done':
                break
            sales = float(input(f"Enter sales for {product_name}: "))
            discount = float(input(f"Enter discount for {product_name} (%): "))
            store_data.append({
                "product_name": product_name,
                "sales": sales,
                "discount": discount
            })
        stores[store_name] = store_data
    return stores

from openpyxl.chart.label import DataLabelList

# STEP 2: Create detailed Excel file with per-store sheets
def create_excel_with_stores(stores, filename="data/store_sales_detailed.xlsx"):
    wb = xl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for store_name, store_data in stores.items():
        sheet = wb.create_sheet(store_name)
        sheet.append(['Product Name', 'Sales', 'Discount (%)', 'Corrected Price', 'Profit'])

        for item in store_data:
            corrected_price = item["sales"] * (1 - item["discount"] / 100)
            profit = corrected_price * 0.30  # 30% profit margin
            sheet.append([item["product_name"], item["sales"], item["discount"], corrected_price, profit])

        categories = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=1)
        sales_values = Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=2)
        corrected_values = Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=4)

        chart_sales = BarChart()
        chart_sales.add_data(sales_values, titles_from_data=True)
        chart_sales.title = f"Sales for {store_name}"
        chart_sales.set_categories(categories)
        chart_sales.display_blanks_as = "zero"

        # Enable data labels
        chart_sales.dLbls = DataLabelList()
        chart_sales.dLbls.showVal = True

        sheet.add_chart(chart_sales, "F2")

        chart_discounted = BarChart()
        chart_discounted.add_data(corrected_values, titles_from_data=True)
        chart_discounted.title = f"Discounted Prices for {store_name}"
        chart_discounted.set_categories(categories)
        chart_discounted.display_blanks_as = "zero"

        # Enable data labels
        chart_discounted.dLbls = DataLabelList()
        chart_discounted.dLbls.showVal = True

        sheet.add_chart(chart_discounted, "F20")

    wb.save(filename)
    print(f"[✓] Detailed Excel file created: {filename}")

from openpyxl.chart.label import DataLabelList

# STEP 3: Create Summary Excel
def create_summary_excel(stores, filename="data/store_sales_summary.xlsx"):
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "All Stores Summary"

    sheet.append(['Store Name', 'Total Sales', 'Total Corrected Price', 'Total Profit'])

    for store_name, store_data in stores.items():
        total_sales = sum([d["sales"] for d in store_data])
        total_corrected = sum([d["sales"] * (1 - d["discount"] / 100) for d in store_data])
        total_profit = total_corrected * 0.30
        sheet.append([store_name, total_sales, total_corrected, total_profit])

    categories = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=1)

    # Total sales chart
    sales_data = Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=2, max_col=2)
    chart_sales = BarChart()
    chart_sales.add_data(sales_data, titles_from_data=True)
    chart_sales.title = "Total Sales for All Stores"
    chart_sales.set_categories(categories)
    chart_sales.display_blanks_as = "zero"

    # Enable data labels
    chart_sales.dLbls = DataLabelList()
    chart_sales.dLbls.showVal = True

    sheet.add_chart(chart_sales, "F2")

    # Total corrected prices chart
    corrected_data = Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=3, max_col=3)
    chart_corrected = BarChart()
    chart_corrected.add_data(corrected_data, titles_from_data=True)
    chart_corrected.title = "Total Discounted Prices for All Stores"
    chart_corrected.set_categories(categories)
    chart_corrected.display_blanks_as = "zero"

    # Enable data labels
    chart_corrected.dLbls = DataLabelList()
    chart_corrected.dLbls.showVal = True

    sheet.add_chart(chart_corrected, "F20")

    wb.save(filename)
    print(f"[✓] Summary Excel file created: {filename}")

# STEP 4: Main function
def main():
    stores = get_store_data()
    create_excel_with_stores(stores)
    create_summary_excel(stores)

if __name__ == "__main__":
    main()
